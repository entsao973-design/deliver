from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from .geocoding import default_geocode_fields
from .import_diagnostics import write_import_log


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def value_after_colon(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    parts = re.split(r"[:：]", text, maxsplit=1)
    return parts[-1].strip()


def normalize_vehicle_no(value: Any) -> str:
    text = value_after_colon(value)
    return re.sub(r"\s+", "", text).upper()


def is_numeric_sequence(value: Any) -> bool:
    text = clean_text(value)
    return bool(re.fullmatch(r"\d+(?:\.0)?", text))


def normalize_sequence(value: Any) -> int:
    return int(float(clean_text(value)))


def parse_delivery_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = value_after_colon(value)
    match = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", text)
    if not match:
        raise ValueError(f"無法解析配送日期: {value!r}")

    year, month, day = (int(part) for part in match.groups())
    return date(year, month, day)


def make_delivery_id(date_folder: str, vehicle_no: str, company: str, invoice_no: str, fallback: str) -> str:
    invoice_key = invoice_no or fallback
    raw = "|".join([date_folder, normalize_vehicle_no(vehicle_no), company, invoice_key])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def import_deliveries(excel_path: str | Path, existing_records: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    excel_path = Path(excel_path)
    existing_by_id = {record["id"]: record for record in existing_records or []}

    write_import_log("excel_importer_start", path=excel_path, bytes=excel_path.stat().st_size if excel_path.exists() else "")
    write_import_log("workbook_load_start", path=excel_path, suffix=excel_path.suffix)
    workbook = openpyxl.load_workbook(
        excel_path,
        read_only=False,
        data_only=True,
        keep_vba=False,
    )
    write_import_log("workbook_load_done", sheets=len(workbook.worksheets), sheet_names=",".join(workbook.sheetnames))
    deliveries: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            if clean_text(sheet["A4"].value) != "序號":
                write_import_log("worksheet_skip", sheet=sheet.title, a4=clean_text(sheet["A4"].value))
                continue

            before_count = len(deliveries)
            write_import_log(
                "worksheet_start",
                sheet=sheet.title,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                merged_ranges=len(sheet.merged_cells.ranges),
            )
            vehicle_no = value_after_colon(sheet["A3"].value)
            vehicle_no_normalized = normalize_vehicle_no(sheet["A3"].value)
            driver = value_after_colon(sheet["C3"].value)
            delivery_date = parse_delivery_date(sheet["H3"].value)
            date_folder = delivery_date.strftime("%Y%m%d")
            merged_sequences = merged_sequence_rows(sheet)
            current_seq: int | None = None
            current_customer = ""
            current_company = ""
            current_address = ""

            for row_index in range(5, sheet.max_row + 1):
                seq_value = merged_sequences.get(row_index, sheet.cell(row_index, 1).value)
                if not is_numeric_sequence(seq_value):
                    continue

                seq = normalize_sequence(seq_value)
                if seq != current_seq:
                    current_seq = seq
                    current_customer = ""
                    current_company = ""
                    current_address = ""

                customer_value = clean_text(sheet.cell(row_index, 2).value)
                address_value = clean_text(sheet.cell(row_index, 3).value)
                company_value = clean_text(sheet.cell(row_index, 9).value)
                if customer_value:
                    current_customer = customer_value
                if address_value:
                    current_address = address_value
                if company_value:
                    current_company = company_value

                customer = current_customer
                address = current_address
                company = current_company
                invoice_no = clean_text(sheet.cell(row_index, 11).value)
                if not invoice_no:
                    continue

                fallback = f"{sheet.title}-R{row_index}"
                delivery_id = make_delivery_id(date_folder, vehicle_no, company, invoice_no, fallback)

                record = {
                    "id": delivery_id,
                    "sheet": sheet.title,
                    "row": row_index,
                    "seq": seq,
                    "vehicle_no": vehicle_no,
                    "vehicle_no_normalized": vehicle_no_normalized,
                    "driver": driver,
                    "delivery_date": delivery_date.isoformat(),
                    "date_folder": date_folder,
                    "customer": customer,
                    "address": address,
                    **default_geocode_fields(address),
                    "company": company,
                    "invoice_no": invoice_no,
                    "status": None,
                    "photo_path": None,
                    "photo_updated_at": None,
                    "updated_at": None,
                }

                previous = existing_by_id.get(delivery_id)
                if previous:
                    for field in ("status", "photo_path", "photo_updated_at", "updated_at"):
                        record[field] = previous.get(field)

                deliveries.append(record)
            write_import_log("worksheet_done", sheet=sheet.title, records=len(deliveries) - before_count)
    finally:
        workbook.close()
        write_import_log("workbook_closed", path=excel_path)

    write_import_log("excel_importer_done", records=len(deliveries))
    return {
        "source_excel": str(excel_path),
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "deliveries": deliveries,
    }


def merged_sequence_rows(sheet) -> dict[int, Any]:
    rows: dict[int, Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row < 5 or not (merged_range.min_col <= 1 <= merged_range.max_col):
            continue

        value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        if not is_numeric_sequence(value):
            continue

        for row_index in range(max(5, merged_range.min_row), merged_range.max_row + 1):
            rows[row_index] = value
    return rows
