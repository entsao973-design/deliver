import json
import os
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

import openpyxl

from delivery_app.excel_importer import import_deliveries
from delivery_app.geocoding import GEOCODE_EMPTY, GEOCODE_FAILED, GEOCODE_PENDING, GEOCODE_SUCCESS
from delivery_app.geocoding import GeocodeResult, StaticGeocoder
from delivery_app.auth import UserStore
from delivery_app.repository import DeliveryRepository


SAMPLE_EXCEL = Path(
    os.environ.get(
        "SAMPLE_EXCEL",
        r"C:\Users\duncan.DUNCAN-PC\Desktop\溫控車配送紀錄 - 2026.06.10北區.中一.xlsm",
    )
)


def build_delivery_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "測試車"
    sheet["A3"] = "車號：TEST-001"
    sheet["C3"] = "物流士：測試司機"
    sheet["H3"] = "配送日期:2026/6/11"
    headers = ["序號", "客戶名稱", "地址", "電話及備註", "收貨時間", "產品簡稱", "數量", "件數", "公司", "類別", "出貨單號"]
    for column, value in enumerate(headers, 1):
        sheet.cell(4, column).value = value

    sheet.merge_cells("A5:A6")
    sheet["A5"] = 1
    sheet["C5"] = "台北市中山區測試路1號"
    sheet["B5"] = "合併客戶"
    sheet["G5"] = "12箱"
    sheet["I5"] = "公司甲"
    sheet["K5"] = "INV-MERGED-1"
    sheet["G6"] = "2箱"
    sheet["I6"] = "公司乙"
    sheet["K6"] = "INV-MERGED-2"
    sheet["A7"] = "備註"
    sheet["K7"] = "IGNORED"
    sheet["A8"] = 2
    sheet["C8"] = "新北市板橋區測試路2號"
    sheet["B8"] = "第二客戶"
    sheet["G8"] = 3
    sheet["I8"] = "公司丙"
    sheet["K8"] = "INV-2"
    sheet["A9"] = None
    sheet["K9"] = "IGNORED-BLANK"
    sheet["A10"] = 3
    sheet["B10"] = "第三客戶"
    sheet["G10"] = ""
    sheet["I10"] = "公司丁"
    sheet["K10"] = "INV-3"
    workbook.save(path)
    workbook.close()


@unittest.skipUnless(SAMPLE_EXCEL.exists(), "sample Excel file not found")
class ExcelImporterTest(unittest.TestCase):
    def test_imports_delivery_sheets(self):
        result = import_deliveries(SAMPLE_EXCEL)
        deliveries = result["deliveries"]

        self.assertEqual(len(deliveries), 16)
        self.assertEqual(deliveries[0]["date_folder"], "20260610")

        vehicles = {record["vehicle_no"] for record in deliveries}
        self.assertIn("RFC-7983", vehicles)
        self.assertIn("RFW-9372", vehicles)
        self.assertIn("RFW-3960", vehicles)

    def test_stops_when_sequence_is_blank(self):
        result = import_deliveries(SAMPLE_EXCEL)
        by_vehicle = {}
        for record in result["deliveries"]:
            by_vehicle.setdefault(record["vehicle_no"], 0)
            by_vehicle[record["vehicle_no"]] += 1

        self.assertEqual(by_vehicle["RFC-7983"], 6)
        self.assertEqual(by_vehicle["RFW-9372"], 6)
        self.assertEqual(by_vehicle["RFW-3960"], 4)

    def test_filters_by_delivery_date(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            repo = DeliveryRepository(
                str(SAMPLE_EXCEL),
                str(Path(temp_dir) / "deliveries.json"),
                str(Path(temp_dir) / "photos"),
            )

            self.assertEqual(
                repo.dates_for_vehicle("RFW-3960"),
                [{"delivery_date": "2026-06-10", "date_folder": "20260610"}],
            )
            self.assertEqual(repo.counts_for_vehicle("RFW-3960", "2026-06-10")["total"], 4)
            self.assertEqual(repo.list_for_vehicle("RFW-3960", include_delivered=True, delivery_date="20260610")[0]["delivery_date"], "2026-06-10")

    def test_reload_preserves_other_dates(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_file = Path(temp_dir) / "deliveries.json"
            repo = DeliveryRepository(str(SAMPLE_EXCEL), str(data_file), str(Path(temp_dir) / "photos"))

            data = json.loads(data_file.read_text(encoding="utf-8"))
            old_record = dict(data["deliveries"][0])
            old_record["id"] = "old-date-record"
            old_record["invoice_no"] = "OLD-DATE-001"
            old_record["delivery_date"] = "2026-06-09"
            old_record["date_folder"] = "20260609"
            data["deliveries"].append(old_record)
            data_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            repo.reload_from_excel()

            dates = repo.dates_for_vehicle(old_record["vehicle_no"])
            self.assertIn({"delivery_date": "2026-06-09", "date_folder": "20260609"}, dates)
            self.assertIn({"delivery_date": "2026-06-10", "date_folder": "20260610"}, dates)

    def test_reimport_updates_undelivered_invoice_and_locks_delivered(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_file = Path(temp_dir) / "deliveries.json"
            repo = DeliveryRepository(str(SAMPLE_EXCEL), str(data_file), str(Path(temp_dir) / "photos"))

            data = json.loads(data_file.read_text(encoding="utf-8"))
            first = data["deliveries"][0]
            first["vehicle_no"] = "OLD-CAR"
            first["vehicle_no_normalized"] = "OLD-CAR"
            data_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            summary = repo.import_excel_file(SAMPLE_EXCEL)
            self.assertGreaterEqual(summary["updated"], 1)
            refreshed = json.loads(data_file.read_text(encoding="utf-8"))
            self.assertEqual(refreshed["deliveries"][0]["vehicle_no"], "RFC-7983")

            refreshed["deliveries"][0]["status"] = "normal"
            refreshed["deliveries"][0]["vehicle_no"] = "LOCKED-CAR"
            data_file.write_text(json.dumps(refreshed, ensure_ascii=False), encoding="utf-8")

            summary = repo.import_excel_file(SAMPLE_EXCEL)
            self.assertGreaterEqual(summary["locked_delivered"], 1)
            locked = json.loads(data_file.read_text(encoding="utf-8"))
            self.assertEqual(locked["deliveries"][0]["vehicle_no"], "LOCKED-CAR")


class ImporterRulesTest(unittest.TestCase):
    def test_reimport_inserts_same_invoice_on_new_date_after_delivery(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            data_file = Path(temp_dir) / "deliveries.json"
            build_delivery_workbook(excel_path)
            DeliveryRepository(str(excel_path), str(data_file), str(Path(temp_dir) / "photos"))

            data = json.loads(data_file.read_text(encoding="utf-8"))
            first = next(record for record in data["deliveries"] if record["invoice_no"] == "INV-MERGED-1")
            first["status"] = "normal"
            first["photo_path"] = "photos/20260611/INV-MERGED-1.JPG"
            first["photo_updated_at"] = "2026-06-11T10:00:00"
            data_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            workbook = openpyxl.load_workbook(excel_path)
            workbook.active["H3"] = "配送日期:2026/6/12"
            workbook.save(excel_path)
            workbook.close()

            repo = DeliveryRepository(None, str(data_file), str(Path(temp_dir) / "photos"))
            summary = repo.import_excel_file(excel_path)
            refreshed = json.loads(data_file.read_text(encoding="utf-8"))["deliveries"]
            matching = [record for record in refreshed if record["invoice_no"] == "INV-MERGED-1"]

            self.assertEqual(summary["inserted"], 1)
            self.assertEqual(len(matching), 2)
            self.assertEqual({record["delivery_date"] for record in matching}, {"2026-06-11", "2026-06-12"})
            delivered = next(record for record in matching if record["delivery_date"] == "2026-06-11")
            new_delivery = next(record for record in matching if record["delivery_date"] == "2026-06-12")
            self.assertEqual(delivered["status"], "normal")
            self.assertIsNone(new_delivery["status"])

    def test_repository_allows_missing_default_excel(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_file = Path(temp_dir) / "deliveries.json"
            repo = DeliveryRepository(None, str(data_file), str(Path(temp_dir) / "photos"))

            self.assertEqual(repo.counts_for_vehicle("TEST-001")["total"], 0)

            excel_path = Path(temp_dir) / "uploaded.xlsx"
            build_delivery_workbook(excel_path)
            summary = repo.import_excel_file(excel_path)

            self.assertEqual(summary["inserted"], 4)
            self.assertEqual(repo.counts_for_vehicle("TEST-001", "2026-06-11")["total"], 4)

    def test_imports_merged_sequence_rows_and_ignores_non_numeric_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)

            result = import_deliveries(excel_path)
            invoices = {record["invoice_no"]: record for record in result["deliveries"]}

            self.assertEqual(set(invoices), {"INV-MERGED-1", "INV-MERGED-2", "INV-2", "INV-3"})
            self.assertEqual(invoices["INV-MERGED-2"]["customer"], "合併客戶")
            self.assertEqual(invoices["INV-MERGED-2"]["company"], "公司乙")
            self.assertEqual(invoices["INV-3"]["seq"], 3)

    def test_imports_address_from_excel_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)

            result = import_deliveries(excel_path)
            invoices = {record["invoice_no"]: record for record in result["deliveries"]}

            self.assertEqual(invoices["INV-MERGED-1"]["address"], "台北市中山區測試路1號")
            self.assertEqual(invoices["INV-MERGED-2"]["address"], "台北市中山區測試路1號")
            self.assertEqual(invoices["INV-2"]["address"], "新北市板橋區測試路2號")

    def test_imports_quantity_from_excel_g_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)

            result = import_deliveries(excel_path)
            invoices = {record["invoice_no"]: record for record in result["deliveries"]}

            self.assertEqual(invoices["INV-MERGED-1"]["quantity"], "12箱")
            self.assertEqual(invoices["INV-MERGED-2"]["quantity"], "2箱")
            self.assertEqual(invoices["INV-2"]["quantity"], "3")
            self.assertEqual(invoices["INV-3"]["quantity"], "")

    def test_import_adds_geocode_defaults(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)

            result = import_deliveries(excel_path)
            invoices = {record["invoice_no"]: record for record in result["deliveries"]}

            self.assertEqual(invoices["INV-MERGED-1"]["geocode_status"], GEOCODE_PENDING)
            self.assertEqual(invoices["INV-MERGED-1"]["normalized_address"], "台北市中山區測試路1號")
            self.assertIsNone(invoices["INV-MERGED-1"]["geocode_lat"])
            self.assertIsNone(invoices["INV-MERGED-1"]["geocode_lng"])
            self.assertEqual(invoices["INV-3"]["geocode_status"], GEOCODE_EMPTY)

    def test_repository_public_records_include_address_and_geocode_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)

            repo = DeliveryRepository(str(excel_path), str(Path(temp_dir) / "deliveries.json"), str(Path(temp_dir) / "photos"))
            deliveries = repo.list_for_vehicle("TEST-001", include_delivered=True, delivery_date="2026-06-11")
            by_invoice = {record["invoice_no"]: record for record in deliveries}

            self.assertEqual(by_invoice["INV-MERGED-1"]["address"], "台北市中山區測試路1號")
            self.assertEqual(by_invoice["INV-MERGED-1"]["quantity"], "12箱")
            self.assertEqual(by_invoice["INV-MERGED-1"]["geocode_status"], GEOCODE_PENDING)
            self.assertIsNone(by_invoice["INV-MERGED-1"]["geocode_lat"])

    def test_repository_preserves_geocode_when_address_is_unchanged(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            data_file = Path(temp_dir) / "deliveries.json"
            build_delivery_workbook(excel_path)
            DeliveryRepository(str(excel_path), str(data_file), str(Path(temp_dir) / "photos"))

            data = json.loads(data_file.read_text(encoding="utf-8"))
            first = next(record for record in data["deliveries"] if record["invoice_no"] == "INV-MERGED-1")
            first["geocode_status"] = GEOCODE_SUCCESS
            first["geocode_lat"] = 25.0478
            first["geocode_lng"] = 121.5319
            first["geocode_provider"] = "static"
            first["geocode_place_id"] = "place-1"
            first["geocode_updated_at"] = "2026-06-17T09:00:00"
            data_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            repo = DeliveryRepository(None, str(data_file), str(Path(temp_dir) / "photos"))
            repo.import_excel_file(excel_path)
            refreshed = json.loads(data_file.read_text(encoding="utf-8"))
            first = next(record for record in refreshed["deliveries"] if record["invoice_no"] == "INV-MERGED-1")

            self.assertEqual(first["geocode_status"], GEOCODE_SUCCESS)
            self.assertEqual(first["geocode_lat"], 25.0478)
            self.assertEqual(first["geocode_lng"], 121.5319)

    def test_repository_resets_geocode_when_address_changes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            data_file = Path(temp_dir) / "deliveries.json"
            build_delivery_workbook(excel_path)
            DeliveryRepository(str(excel_path), str(data_file), str(Path(temp_dir) / "photos"))

            data = json.loads(data_file.read_text(encoding="utf-8"))
            first = next(record for record in data["deliveries"] if record["invoice_no"] == "INV-MERGED-1")
            first["geocode_status"] = GEOCODE_SUCCESS
            first["geocode_lat"] = 25.0478
            first["geocode_lng"] = 121.5319
            data_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            workbook = openpyxl.load_workbook(excel_path)
            workbook.active["C5"] = "台北市中山區新地址99號"
            workbook.save(excel_path)
            workbook.close()

            repo = DeliveryRepository(None, str(data_file), str(Path(temp_dir) / "photos"))
            repo.import_excel_file(excel_path)
            refreshed = json.loads(data_file.read_text(encoding="utf-8"))
            first = next(record for record in refreshed["deliveries"] if record["invoice_no"] == "INV-MERGED-1")

            self.assertEqual(first["address"], "台北市中山區新地址99號")
            self.assertEqual(first["geocode_status"], GEOCODE_PENDING)
            self.assertIsNone(first["geocode_lat"])
            self.assertIsNone(first["geocode_lng"])

    def test_repository_geocodes_pending_addresses_and_reuses_cache(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            data_file = Path(temp_dir) / "deliveries.json"
            build_delivery_workbook(excel_path)
            geocoder = StaticGeocoder({
                "台北市中山區測試路1號": GeocodeResult.success(
                    lat=25.0478,
                    lng=121.5319,
                    provider="static",
                    place_id="place-1",
                ),
            })
            repo = DeliveryRepository(
                str(excel_path),
                str(data_file),
                str(Path(temp_dir) / "photos"),
                geocoder=geocoder,
            )

            summary = repo.geocode_pending()
            data = json.loads(data_file.read_text(encoding="utf-8"))
            by_invoice = {record["invoice_no"]: record for record in data["deliveries"]}

            self.assertEqual(summary["success"], 2)
            self.assertEqual(by_invoice["INV-MERGED-1"]["geocode_status"], GEOCODE_SUCCESS)
            self.assertEqual(by_invoice["INV-MERGED-2"]["geocode_status"], GEOCODE_SUCCESS)
            self.assertEqual(by_invoice["INV-MERGED-1"]["geocode_lat"], 25.0478)
            self.assertIn("台北市中山區測試路1號", data["geocode_cache"])

    def test_repository_does_not_cache_failed_geocoding_results(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            data_file = Path(temp_dir) / "deliveries.json"
            build_delivery_workbook(excel_path)
            geocoder = StaticGeocoder({
                "台北市中山區測試路1號": GeocodeResult(
                    status=GEOCODE_FAILED,
                    normalized_address="台北市中山區測試路1號",
                    provider="static",
                    error_message="temporary failure",
                ),
            })
            repo = DeliveryRepository(
                str(excel_path),
                str(data_file),
                str(Path(temp_dir) / "photos"),
                geocoder=geocoder,
            )

            summary = repo.geocode_pending()
            data = json.loads(data_file.read_text(encoding="utf-8"))

            self.assertEqual(summary["failed"], 2)
            self.assertNotIn("台北市中山區測試路1號", data["geocode_cache"])

    def test_vehicle_options_use_latest_delivery_date(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            data_file = Path(temp_dir) / "deliveries.json"
            build_delivery_workbook(excel_path)
            DeliveryRepository(str(excel_path), str(data_file), str(Path(temp_dir) / "photos"))

            data = json.loads(data_file.read_text(encoding="utf-8"))
            older_record = dict(data["deliveries"][0])
            older_record["id"] = "older-vehicle"
            older_record["vehicle_no"] = "OLD-CAR"
            older_record["vehicle_no_normalized"] = "OLD-CAR"
            older_record["delivery_date"] = "2026-06-10"
            older_record["date_folder"] = "20260610"
            data["deliveries"].append(older_record)
            data_file.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

            repo = DeliveryRepository(None, str(data_file), str(Path(temp_dir) / "photos"))

            self.assertEqual(
                repo.vehicles_for_latest_date(),
                {
                    "delivery_date": "2026-06-11",
                    "vehicles": ["TEST-001"],
                    "vehicle_options": [{"vehicle_no": "TEST-001", "driver": "測試司機"}],
                },
            )

    def test_filter_options_use_date_range_and_deleted_state(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            data_file = Path(temp_dir) / "deliveries.json"
            data_file.write_text(
                json.dumps(
                    {
                        "deliveries": [
                            make_delivery_record("old", "2026-06-10", "OldCo", "OldDriver"),
                            make_delivery_record("in-range", "2026-06-12", "RangeCo", "RangeDriver"),
                            make_delivery_record("new", "2026-06-14", "NewCo", "NewDriver"),
                            make_delivery_record(
                                "deleted",
                                "2026-06-12",
                                "DeletedCo",
                                "DeletedDriver",
                                deleted_at="2026-06-13T10:00:00",
                            ),
                        ],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            repo = DeliveryRepository(None, str(data_file), str(Path(temp_dir) / "photos"))

            self.assertEqual(
                repo.filter_options(start_date="2026-06-11", end_date="2026-06-13"),
                {
                    "dates": ["2026-06-12"],
                    "companies": ["RangeCo"],
                    "drivers": ["RangeDriver"],
                },
            )
            self.assertEqual(
                repo.filter_options(start_date="2026-06-11", end_date="2026-06-13", deleted=True),
                {
                    "dates": ["2026-06-12"],
                    "companies": ["DeletedCo"],
                    "drivers": ["DeletedDriver"],
                },
            )

    def test_photo_path_uses_company_folder_and_status_suffix(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)
            repo = DeliveryRepository(str(excel_path), str(Path(temp_dir) / "deliveries.json"), str(Path(temp_dir) / "photos"))
            delivery = repo.list_for_vehicle("TEST-001", include_delivered=True, delivery_date="2026-06-11")[0]

            repo.update_photo(delivery["id"], "abnormal", "data:image/jpeg;base64,dGVzdA==")
            path = repo.photo_path_for(delivery["id"])

            self.assertIsNotNone(path)
            self.assertEqual(path.name, f"{delivery['invoice_no']}_異常.JPG")
            self.assertEqual(path.parent.name, delivery["company"])
            self.assertEqual(path.parent.parent.name, "20260611")

    def test_update_photo_uses_captured_at_for_photo_time(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)
            repo = DeliveryRepository(str(excel_path), str(Path(temp_dir) / "deliveries.json"), str(Path(temp_dir) / "photos"))
            delivery = repo.list_for_vehicle("TEST-001", include_delivered=True, delivery_date="2026-06-11")[0]

            updated = repo.update_photo(
                delivery["id"],
                "normal",
                "data:image/jpeg;base64,dGVzdA==",
                captured_at="2026-06-14T09:30:00",
            )

            self.assertEqual(updated["photo_updated_at"], "2026-06-14T09:30:00")

    def test_archive_photos_creates_company_zip(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "merged.xlsx"
            build_delivery_workbook(excel_path)
            repo = DeliveryRepository(str(excel_path), str(Path(temp_dir) / "deliveries.json"), str(Path(temp_dir) / "photos"))
            delivery = repo.list_for_vehicle("TEST-001", include_delivered=True, delivery_date="2026-06-11")[0]
            repo.update_photo(delivery["id"], "normal", "data:image/jpeg;base64,dGVzdA==")

            archives = repo.archive_photos("2026-06-11")
            archive = next(item for item in archives if item["company"] == delivery["company"])
            archive_path = repo.archive_path_for(archive["name"])

            self.assertIsNotNone(archive_path)
            with zipfile.ZipFile(archive_path) as zip_file:
                self.assertIn(f"{delivery['invoice_no']}.JPG", zip_file.namelist())

    def test_list_archives_returns_existing_zips_for_selected_date(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            archive_root = root / "archives"
            archive_root.mkdir()
            (archive_root / "20260611_Alpha.zip").write_bytes(b"alpha")
            (archive_root / "20260611_Beta.ZIP").write_bytes(b"beta")
            (archive_root / "20260612_Other.zip").write_bytes(b"other")
            (archive_root / "20260611_notes.txt").write_text("ignore", encoding="utf-8")
            repo = DeliveryRepository(
                None,
                str(root / "deliveries.json"),
                str(root / "photos"),
                str(archive_root),
            )

            archives = repo.list_archives("2026-06-11")

            self.assertEqual(
                archives,
                [
                    {
                        "name": "20260611_Alpha.zip",
                        "company": "Alpha",
                        "date_folder": "20260611",
                        "size": 5,
                    },
                    {
                        "name": "20260611_Beta.ZIP",
                        "company": "Beta",
                        "date_folder": "20260611",
                        "size": 4,
                    },
                ],
            )

    def test_cleanup_delivery_history_removes_all_records_and_files_in_inclusive_range(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_file = root / "deliveries.json"
            photo_root = root / "photos"
            archive_root = root / "archives"
            records = [
                make_delivery_record("before", "2026-06-09", "Before", "Driver"),
                make_delivery_record("open", "2026-06-10", "Open", "Driver"),
                {**make_delivery_record("done", "2026-06-11", "Done", "Driver"), "status": "normal"},
                {
                    **make_delivery_record(
                        "deleted",
                        "2026-06-12",
                        "Deleted",
                        "Driver",
                        "2026-06-13T10:00:00",
                    ),
                    "status": "abnormal",
                },
                make_delivery_record("after", "2026-06-13", "After", "Driver"),
            ]
            data_file.write_text(json.dumps({"deliveries": records}, ensure_ascii=False), encoding="utf-8")
            for folder in (
                "20260609",
                "20260610",
                "20260611",
                "20260612",
                "20260613",
                "20261340",
                "notes",
            ):
                (photo_root / folder).mkdir(parents=True)
                (photo_root / folder / "file.jpg").write_bytes(b"photo")
            archive_root.mkdir()
            for filename in (
                "20260609_Before.zip",
                "20260610_Open.zip",
                "20260612_Deleted.ZIP",
                "20260613_After.zip",
                "20260610NoUnderscore.zip",
                "20261340_Invalid.zip",
                "unmatched.zip",
            ):
                (archive_root / filename).write_bytes(b"zip")

            repo = DeliveryRepository(None, str(data_file), str(photo_root), str(archive_root))
            summary = repo.cleanup_delivery_history("2026-06-10", "2026-06-12")

            saved_ids = [
                item["id"]
                for item in json.loads(data_file.read_text(encoding="utf-8"))["deliveries"]
            ]
            self.assertEqual(saved_ids, ["before", "after"])
            self.assertEqual(
                summary,
                {
                    "deleted_records": 3,
                    "deleted_photo_date_folders": 3,
                    "deleted_archives": 2,
                },
            )
            self.assertFalse((photo_root / "20260610").exists())
            self.assertFalse((photo_root / "20260611").exists())
            self.assertFalse((photo_root / "20260612").exists())
            self.assertTrue((photo_root / "20260609").exists())
            self.assertTrue((photo_root / "20260613").exists())
            self.assertTrue((photo_root / "20261340").exists())
            self.assertTrue((photo_root / "notes").exists())
            self.assertTrue((archive_root / "20260609_Before.zip").exists())
            self.assertTrue((archive_root / "20260613_After.zip").exists())
            self.assertTrue((archive_root / "20260610NoUnderscore.zip").exists())
            self.assertTrue((archive_root / "20261340_Invalid.zip").exists())
            self.assertTrue((archive_root / "unmatched.zip").exists())

            self.assertEqual(
                repo.cleanup_delivery_history("2026-06-10", "2026-06-12"),
                {
                    "deleted_records": 0,
                    "deleted_photo_date_folders": 0,
                    "deleted_archives": 0,
                },
            )

    def test_cleanup_delivery_history_rejects_invalid_date_range(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            repo = DeliveryRepository(
                None,
                str(root / "deliveries.json"),
                str(root / "photos"),
                str(root / "archives"),
            )

            with self.assertRaisesRegex(ValueError, "日期格式"):
                repo.cleanup_delivery_history("2026-02-30", "2026-03-01")
            with self.assertRaisesRegex(ValueError, "日期格式"):
                repo.cleanup_delivery_history("2026-6-01", "2026-06-02")
            with self.assertRaisesRegex(ValueError, "開始日期"):
                repo.cleanup_delivery_history("2026-03-02", "2026-03-01")

    def test_cleanup_delivery_history_reports_incomplete_file_cleanup_after_records_are_removed(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_file = root / "deliveries.json"
            photo_root = root / "photos"
            archive_root = root / "archives"
            data_file.write_text(
                json.dumps(
                    {
                        "deliveries": [
                            make_delivery_record("remove", "2026-06-10", "Company", "Driver"),
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (photo_root / "20260610").mkdir(parents=True)
            archive_root.mkdir()
            repo = DeliveryRepository(None, str(data_file), str(photo_root), str(archive_root))

            with patch("delivery_app.repository.shutil.rmtree", side_effect=PermissionError("locked")):
                with self.assertRaisesRegex(RuntimeError, "清理未完整完成"):
                    repo.cleanup_delivery_history("2026-06-10", "2026-06-10")

            saved = json.loads(data_file.read_text(encoding="utf-8"))
            self.assertEqual(saved["deliveries"], [])

    def test_delete_delivery_moves_pending_records_to_deleted_area(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_file = root / "deliveries.json"
            records = [
                make_delivery_record("pending", "2026-06-10", "Open", "Driver"),
            ]
            data_file.write_text(json.dumps({"deliveries": records}, ensure_ascii=False), encoding="utf-8")
            repo = DeliveryRepository(None, str(data_file), str(root / "photos"), str(root / "archives"))

            result = repo.delete_delivery("pending", "admin")

            saved = json.loads(data_file.read_text(encoding="utf-8"))["deliveries"][0]
            self.assertEqual(result["mode"], "archived")
            self.assertEqual(saved["id"], "pending")
            self.assertIsNotNone(saved["deleted_at"])
            self.assertEqual(saved["deleted_by"], "admin")
            self.assertEqual([item["id"] for item in repo.list_admin_deliveries(deleted=True)], ["pending"])

    def test_bulk_delete_and_permanent_delete_use_selected_ids_only(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_file = root / "deliveries.json"
            photo_root = root / "photos"
            archive_root = root / "archives"
            selected_photo = photo_root / "20260610" / "Done" / "INV-selected.JPG"
            kept_photo = photo_root / "20260610" / "Done" / "INV-kept.JPG"
            selected_photo.parent.mkdir(parents=True)
            selected_photo.write_bytes(b"selected")
            kept_photo.write_bytes(b"kept")
            archive_root.mkdir()
            selected_archive = archive_root / "20260610_SelectedCo.zip"
            kept_archive = archive_root / "20260610_KeptCo.zip"
            selected_archive.write_bytes(b"selected zip")
            kept_archive.write_bytes(b"kept zip")
            records = [
                make_delivery_record("pending", "2026-06-10", "Open", "Driver"),
                {**make_delivery_record("done", "2026-06-10", "Done", "Driver"), "status": "normal"},
                {
                    **make_delivery_record("selected", "2026-06-10", "SelectedCo", "Driver", "2026-06-11T10:00:00"),
                    "status": "normal",
                    "photo_path": str(selected_photo),
                },
                {
                    **make_delivery_record("kept", "2026-06-10", "KeptCo", "Driver", "2026-06-11T10:00:00"),
                    "status": "normal",
                    "photo_path": str(kept_photo),
                },
            ]
            data_file.write_text(json.dumps({"deliveries": records}, ensure_ascii=False), encoding="utf-8")
            repo = DeliveryRepository(None, str(data_file), str(photo_root), str(archive_root))

            delete_summary = repo.delete_deliveries(["pending", "done"], "admin")
            permanent_summary = repo.permanently_delete_deliveries(["selected"])

            saved = json.loads(data_file.read_text(encoding="utf-8"))["deliveries"]
            by_id = {item["id"]: item for item in saved}
            self.assertEqual(delete_summary, {"deleted_records": 2})
            self.assertEqual(permanent_summary, {"deleted_records": 1})
            self.assertEqual(set(by_id), {"pending", "done", "kept"})
            self.assertIsNotNone(by_id["pending"]["deleted_at"])
            self.assertIsNotNone(by_id["done"]["deleted_at"])
            self.assertIsNotNone(by_id["kept"]["deleted_at"])
            self.assertFalse(selected_photo.exists())
            self.assertTrue(kept_photo.exists())
            self.assertFalse(selected_archive.exists())
            self.assertTrue(kept_archive.exists())

    def test_permanent_delete_delivery_removes_matching_archive_zip(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            data_file = root / "deliveries.json"
            photo_root = root / "photos"
            archive_root = root / "archives"
            archive_root.mkdir()
            selected_archive = archive_root / "20260610_SelectedCo.zip"
            kept_archive = archive_root / "20260610_KeptCo.zip"
            selected_archive.write_bytes(b"selected zip")
            kept_archive.write_bytes(b"kept zip")
            records = [
                make_delivery_record("selected", "2026-06-10", "SelectedCo", "Driver", "2026-06-11T10:00:00"),
                make_delivery_record("kept", "2026-06-10", "KeptCo", "Driver", "2026-06-11T10:00:00"),
            ]
            data_file.write_text(json.dumps({"deliveries": records}, ensure_ascii=False), encoding="utf-8")
            repo = DeliveryRepository(None, str(data_file), str(photo_root), str(archive_root))

            repo.permanently_delete_delivery("selected")

            saved_ids = [item["id"] for item in json.loads(data_file.read_text(encoding="utf-8"))["deliveries"]]
            self.assertEqual(saved_ids, ["kept"])
            self.assertFalse(selected_archive.exists())
            self.assertTrue(kept_archive.exists())


class UserStoreTest(unittest.TestCase):
    def test_locks_after_five_failed_logins(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = UserStore(str(Path(temp_dir) / "users.json"), [{"username": "driver", "password": "1234"}])

            for index in range(4):
                ok, user, message = store.authenticate("driver", "wrong")
                self.assertFalse(ok)
                self.assertIn(f"已失敗 {index + 1} 次", message)

            ok, user, message = store.authenticate("driver", "wrong")
            self.assertFalse(ok)
            self.assertIn("鎖定 10 分鐘", message)

            ok, user, message = store.authenticate("driver", "1234")
            self.assertFalse(ok)
            self.assertIn("分鐘後再試", message)


def make_delivery_record(
    record_id: str,
    delivery_date: str,
    company: str,
    driver: str,
    deleted_at: str | None = None,
) -> dict[str, object]:
    return {
        "id": record_id,
        "seq": 1,
        "vehicle_no": "TEST-001",
        "vehicle_no_normalized": "TEST-001",
        "driver": driver,
        "delivery_date": delivery_date,
        "date_folder": delivery_date.replace("-", ""),
        "customer": f"Customer {record_id}",
        "company": company,
        "invoice_no": f"INV-{record_id}",
        "status": None,
        "photo_path": None,
        "photo_updated_at": None,
        "updated_at": None,
        "deleted_at": deleted_at,
        "deleted_by": "admin" if deleted_at else None,
    }


if __name__ == "__main__":
    unittest.main()
